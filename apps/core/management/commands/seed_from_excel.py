"""
Management command: seed_from_excel
------------------------------------
Reads facebook/seed_data.xlsx (produced by facebook/extract_to_excel.py)
and loads users + listings + images into the database.

Usage:
    python manage.py seed_from_excel --sheet users

    python manage.py seed_from_excel
    python manage.py seed_from_excel --excel ../facebook/seed_data.xlsx
    python manage.py seed_from_excel --sheet PropertyListing --skip-images
    python manage.py seed_from_excel --dry-run

    python manage.py seed_from_excel --limit 50 --sheet PropertyListing
    python manage.py seed_from_excel --limit 10 --skip-images --dry-run

"""

import sys
import json as _json
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

try:
    import openpyxl
except ImportError:
    sys.exit("Install openpyxl:  pip install openpyxl")


# ── Helpers ────────────────────────────────────────────────────────────────────

def _bool(val) -> bool:
    if isinstance(val, bool):
        return val
    return str(val).strip().upper() in ('TRUE', '1', 'YES')


def _int(val, default=0) -> int:
    try:
        return int(float(str(val).replace(',', '')))
    except (TypeError, ValueError):
        return default


def _decimal(val):
    from decimal import Decimal, InvalidOperation
    try:
        return Decimal(str(val).replace(',', ''))
    except InvalidOperation:
        return Decimal('0')


def _str(val) -> str:
    return str(val).strip() if val is not None else ''


def _read_sheet(ws) -> list[dict]:
    # iter_rows reads the file sequentially (one pass); ws.cell(row, col) in
    # read_only mode re-parses XML from the start for every access — very slow.
    rows_iter = ws.iter_rows(values_only=True)
    headers = list(next(rows_iter, []))
    if not headers:
        return []
    rows = []
    for values in rows_iter:
        row = dict(zip(headers, values))
        if any(v is not None and str(v).strip() for v in row.values()):
            rows.append(row)
    return rows


def _get_or_create_user(phone: str, row: dict):
    """Find existing user by phone or email, or create via the proper manager."""
    from apps.accounts.models import User
    email = _str(row.get('email')) or f"{phone}@barmha.phone"

    user = User.objects.filter(phone=phone).first()
    if user:
        return user, False
    user = User.objects.filter(email=email).first()
    if user:
        return user, False

    first = _str(row.get('first_name'))
    last  = _str(row.get('last_name'))
    user = User.objects.create_user(
        email      = email,
        password   = phone,
        phone      = phone,
        first_name = first,
        last_name  = last,
        user_type  = 'individual',
        language   = 'ar',
    )
    return user, True


def _build_location_cache() -> dict:
    """Return {governorate_slug: Location} — one query instead of N."""
    from apps.core.models import Location
    cache = {}
    for loc in Location.objects.select_related('governorate').all():
        slug = loc.governorate.slug
        if slug not in cache:
            cache[slug] = loc
    return cache


def _upload_image(path_str: str):
    from django.core.files.uploadedfile import SimpleUploadedFile
    p = Path(path_str)
    if not p.exists():
        return None
    suffix = p.suffix.lower().lstrip('.')
    mime = {
        'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
        'png': 'image/png',  'webp': 'image/webp',
    }.get(suffix, 'image/jpeg')
    with open(p, 'rb') as f:
        data = f.read()
    return SimpleUploadedFile(p.name, data, content_type=mime)


def _progress(stdout, label: str, idx: int, total: int):
    if idx == 1 or idx % 25 == 0 or idx == total:
        stdout.write(f"  [{label}] row {idx}/{total}")


# Rows per bulk_create INSERT for PropertyListing / VehicleListing.
BULK_SIZE = 200


# ── Users ──────────────────────────────────────────────────────────────────────

USER_BATCH_SIZE = 25


def seed_users(rows: list[dict], dry_run: bool, stdout) -> dict[str, object]:
    """Create users from the users sheet. Returns phone → User mapping."""
    from apps.accounts.models import User

    user_map: dict[str, object] = {}
    created = skipped = errors = 0
    total = len(rows)
    stdout.write(f"  {total} user rows to process ...")

    if dry_run:
        for idx, row in enumerate(rows, 1):
            phone = _str(row.get('phone'))
            if not phone:
                continue
            user = (
                User.objects.filter(phone=phone).first()
                or User.objects.filter(
                    email=_str(row.get('email')) or f"{phone}@barmha.phone"
                ).first()
            )
            user_map[phone] = user
            if user:
                skipped += 1
            else:
                created += 1
            _progress(stdout, "users", idx, total)
        stdout.write(f"  Users done: {created} created, {skipped} already exist, {errors} errors")
        return user_map

    # Collect rows into batches of USER_BATCH_SIZE, commit each batch atomically.
    batch: list[dict] = []

    def _flush(batch: list[dict]) -> tuple[int, int, int]:
        b_created = b_skipped = b_errors = 0
        with transaction.atomic():
            for row in batch:
                phone = _str(row.get('phone'))
                if not phone:
                    continue
                try:
                    user, was_created = _get_or_create_user(phone, row)
                    user_map[phone] = user
                    if was_created:
                        b_created += 1
                    else:
                        b_skipped += 1
                except Exception as exc:
                    b_errors += 1
                    stdout.write(f"  [user error] phone={phone}: {exc}")
        return b_created, b_skipped, b_errors

    for idx, row in enumerate(rows, 1):
        phone = _str(row.get('phone'))
        if not phone:
            continue
        batch.append(row)

        if len(batch) >= USER_BATCH_SIZE:
            bc, bs, be = _flush(batch)
            created += bc; skipped += bs; errors += be
            stdout.write(f"  [users] committed batch → {created} created so far")
            batch = []

        _progress(stdout, "users", idx, total)

    # Flush remaining rows that didn't fill a full batch.
    if batch:
        bc, bs, be = _flush(batch)
        created += bc; skipped += bs; errors += be
        stdout.write(f"  [users] committed final batch → {created} created so far")

    stdout.write(f"  Users done: {created} created, {skipped} already exist, {errors} errors")
    return user_map


# ── PropertyListing — bulk_create ─────────────────────────────────────────────

def seed_properties(rows: list[dict], user_map: dict, loc_cache: dict,
                    dry_run: bool, skip_images: bool, stdout) -> int:
    if not user_map:
        stdout.write("  [WARN] user_map is empty — all property rows will be skipped. Seed users first.")
        return 0

    total = len(rows)
    stdout.write(f"  {total} property rows to process ...")

    if dry_run:
        would_create = sum(1 for r in rows if user_map.get(_str(r.get('phone'))))
        stdout.write(f"  PropertyListing done: {would_create} would create, {total - would_create} no-user")
        return would_create

    from apps.properties.models import PropertyListing, PropertyImage

    # Build phase — construct unsaved model instances; zero DB calls here.
    to_create: list[tuple] = []   # (PropertyListing, images_str)
    skipped = build_errors = 0

    for row in rows:
        user = user_map.get(_str(row.get('phone')))
        if not user:
            skipped += 1
            continue
        try:
            post_id = _str(row.get('source_post_id'))
            loc     = loc_cache.get(_str(row.get('governorate_slug')))
            amenities = []
            raw = _str(row.get('amenities', '[]'))
            if raw:
                try:
                    amenities = _json.loads(raw)
                except Exception:
                    pass

            to_create.append((
                PropertyListing(
                    posted_by      = user,
                    listing_intent = _str(row.get('listing_intent')) or 'offer',
                    title          = _str(row.get('title')) or f'إعلان {post_id}',
                    description    = _str(row.get('description')),
                    purpose        = _str(row.get('purpose')) or 'for_sale',
                    category       = _str(row.get('category')) or 'residential',
                    property_type  = _str(row.get('property_type')) or 'apartment',
                    address_detail = _str(row.get('address_detail')),
                    location       = loc,
                    price          = _decimal(row.get('price') or 0),
                    currency       = _str(row.get('currency')) or 'SYP',
                    negotiable     = _bool(row.get('negotiable')),
                    hide_price     = _bool(row.get('hide_price')) or not row.get('price'),
                    bedrooms       = _str(row.get('bedrooms')),
                    bathrooms      = _int(row.get('bathrooms')) or None,
                    area_sqm       = _decimal(row.get('area_sqm')) if row.get('area_sqm') else None,
                    floor_number   = _int(row.get('floor_number')) if row.get('floor_number') else None,
                    furnishing     = _str(row.get('furnishing')),
                    amenities      = amenities,
                    views_count    = _int(row.get('views_count')),
                    is_active      = True,
                ),
                _str(row.get('images', '')),
            ))
        except Exception as exc:
            build_errors += 1
            stdout.write(f"  [property build] {row.get('source_post_id')}: {exc}")

    stdout.write(f"  Built {len(to_create)} instances ({skipped} skipped, {build_errors} build errors)")

    # Bulk-create phase — one INSERT per BULK_SIZE rows instead of one per row.
    all_objs = [t[0] for t in to_create]
    all_imgs = [t[1] for t in to_create]
    created = img_errors = 0

    for i in range(0, len(all_objs), BULK_SIZE):
        chunk      = all_objs[i : i + BULK_SIZE]
        chunk_imgs = all_imgs[i : i + BULK_SIZE]

        with transaction.atomic():
            saved = PropertyListing.objects.bulk_create(chunk)
        created += len(saved)
        stdout.write(f"  [property] bulk_create → {created}/{len(all_objs)} saved")

        # Images are uploaded AFTER the transaction commits so a bad upload
        # cannot roll back the already-persisted listings.
        if not skip_images:
            for listing, imgs_str in zip(saved, chunk_imgs):
                if not imgs_str:
                    continue
                for order, path in enumerate(
                    p.strip() for p in imgs_str.split(';') if p.strip()
                ):
                    try:
                        img_file = _upload_image(path)
                        if img_file:
                            pi = PropertyImage(listing=listing, order=order, is_primary=(order == 0))
                            pi.image.save(Path(path).name, img_file, save=True)
                    except Exception:
                        img_errors += 1

    stdout.write(
        f"  PropertyListing done: {created} created, {skipped} no-user, "
        f"{build_errors} build errors, {img_errors} image errors"
    )
    return created


# ── VehicleListing — bulk_create ──────────────────────────────────────────────

def seed_vehicles(rows: list[dict], user_map: dict, loc_cache: dict,
                  dry_run: bool, skip_images: bool, stdout) -> int:
    if not user_map:
        stdout.write("  [WARN] user_map is empty — all vehicle rows will be skipped. Seed users first.")
        return 0

    total = len(rows)
    stdout.write(f"  {total} vehicle rows to process ...")

    if dry_run:
        would_create = sum(1 for r in rows if user_map.get(_str(r.get('phone'))))
        stdout.write(f"  VehicleListing done: {would_create} would create, {total - would_create} no-user")
        return would_create

    from apps.vehicles.models import VehicleListing, VehicleImage, VehicleMake

    # Pre-load all makes into a dict to avoid one DB hit per vehicle row.
    make_cache: dict[str, object] = {m.name.lower(): m for m in VehicleMake.objects.all()}

    def _find_make(name: str):
        if not name:
            return None
        nl = name.lower()
        if nl in make_cache:
            return make_cache[nl]
        for k, v in make_cache.items():
            if nl in k or k in nl:
                return v
        return None

    to_create: list[tuple] = []
    skipped = build_errors = 0

    for row in rows:
        user = user_map.get(_str(row.get('phone')))
        if not user:
            skipped += 1
            continue
        try:
            post_id = _str(row.get('source_post_id'))
            loc     = loc_cache.get(_str(row.get('governorate_slug')))

            to_create.append((
                VehicleListing(
                    posted_by      = user,
                    listing_intent = _str(row.get('listing_intent')) or 'offer',
                    listing_type   = _str(row.get('listing_type')) or 'sale',
                    category       = _str(row.get('category')) or 'car',
                    title          = _str(row.get('title')) or f'إعلان {post_id}',
                    description    = _str(row.get('description')),
                    make           = _find_make(_str(row.get('make'))),
                    year           = _int(row.get('year')) or None,
                    condition      = _str(row.get('condition')) or 'used',
                    price          = _decimal(row.get('price') or 0),
                    currency       = _str(row.get('currency')) or 'SYP',
                    negotiable     = _bool(row.get('negotiable')),
                    location       = loc,
                    views_count    = _int(row.get('views_count')),
                    is_active      = True,
                ),
                _str(row.get('images', '')),
            ))
        except Exception as exc:
            build_errors += 1
            stdout.write(f"  [vehicle build] {row.get('source_post_id')}: {exc}")

    stdout.write(f"  Built {len(to_create)} instances ({skipped} skipped, {build_errors} build errors)")

    all_objs = [t[0] for t in to_create]
    all_imgs = [t[1] for t in to_create]
    created = img_errors = 0

    for i in range(0, len(all_objs), BULK_SIZE):
        chunk      = all_objs[i : i + BULK_SIZE]
        chunk_imgs = all_imgs[i : i + BULK_SIZE]

        with transaction.atomic():
            saved = VehicleListing.objects.bulk_create(chunk)
        created += len(saved)
        stdout.write(f"  [vehicle] bulk_create → {created}/{len(all_objs)} saved")

        if not skip_images:
            for listing, imgs_str in zip(saved, chunk_imgs):
                if not imgs_str:
                    continue
                for order, path in enumerate(
                    p.strip() for p in imgs_str.split(';') if p.strip()
                ):
                    try:
                        img_file = _upload_image(path)
                        if img_file:
                            vi = VehicleImage(listing=listing, order=order, is_primary=(order == 0))
                            vi.image.save(Path(path).name, img_file, save=True)
                    except Exception:
                        img_errors += 1

    stdout.write(
        f"  VehicleListing done: {created} created, {skipped} no-user, "
        f"{build_errors} build errors, {img_errors} image errors"
    )
    return created


# ── ClassifiedListing — one-by-one (django-parler translated fields) ───────────

def _save_classified_row(row: dict, user_map: dict, loc_cache: dict, skip_images: bool) -> int:
    """Returns number of image errors."""
    from apps.classifieds.models import ClassifiedListing, ClassifiedImage

    phone   = _str(row.get('phone'))
    post_id = _str(row.get('source_post_id'))
    user    = user_map.get(phone)
    if not user:
        raise ValueError(f"no user for phone={phone}")

    loc = loc_cache.get(_str(row.get('governorate_slug')))
    listing = ClassifiedListing(
        seller       = user,
        listing_type = _str(row.get('listing_type')) or 'sell',
        condition    = _str(row.get('condition')) or 'used_good',
        price        = _decimal(row.get('price') or 0),
        currency     = _str(row.get('currency')) or 'SYP',
        negotiable   = _bool(row.get('negotiable')),
        location     = loc,
        views_count  = _int(row.get('views_count')),
        is_active    = True,
    )
    listing.save()
    listing.set_current_language('ar')
    listing.title       = _str(row.get('title')) or f'إعلان {post_id}'
    listing.description = _str(row.get('description'))
    listing.save()

    img_errors = 0
    if not skip_images:
        images_str = _str(row.get('images', ''))
        for order, path in enumerate(
            p.strip() for p in images_str.split(';') if p.strip()
        ):
            try:
                img_file = _upload_image(path)
                if img_file:
                    ci = ClassifiedImage(listing=listing, order=order, is_primary=(order == 0))
                    ci.image.save(Path(path).name, img_file, save=True)
            except Exception:
                img_errors += 1
    return img_errors


def seed_classifieds(rows: list[dict], user_map: dict, loc_cache: dict,
                     dry_run: bool, skip_images: bool, stdout) -> int:
    if not user_map:
        stdout.write("  [WARN] user_map is empty — all classified rows will be skipped. Seed users first.")
        return 0

    total = len(rows)
    stdout.write(f"  {total} classified rows to process ...")

    if dry_run:
        would_create = sum(1 for r in rows if user_map.get(_str(r.get('phone'))))
        stdout.write(f"  ClassifiedListing done: {would_create} would create, {total - would_create} no-user")
        return would_create

    created = skipped = errors = img_errors = 0
    for idx, row in enumerate(rows, 1):
        if not user_map.get(_str(row.get('phone'))):
            skipped += 1
            continue
        post_id = _str(row.get('source_post_id'))
        try:
            with transaction.atomic():
                img_errors += _save_classified_row(row, user_map, loc_cache, skip_images)
            created += 1
        except Exception as exc:
            errors += 1
            stdout.write(f"  [classified error] {post_id}: {exc}")
        _progress(stdout, "classified", idx, total)

    stdout.write(
        f"  ClassifiedListing done: {created} created, {skipped} no-user, "
        f"{errors} errors, {img_errors} image errors"
    )
    return created


# ── JobListing — one-by-one (django-parler translated fields) ──────────────────

def _save_job_row(row: dict, user_map: dict, loc_cache: dict) -> None:
    from apps.jobs.models import JobListing

    phone   = _str(row.get('phone'))
    post_id = _str(row.get('source_post_id'))
    user    = user_map.get(phone)
    if not user:
        raise ValueError(f"no user for phone={phone}")

    loc = loc_cache.get(_str(row.get('governorate_slug')))
    listing = JobListing(
        posted_by   = user,
        job_type    = _str(row.get('job_type')) or 'full_time',
        experience  = _str(row.get('experience')),
        currency    = _str(row.get('currency')) or 'SYP',
        min_salary  = _decimal(row.get('min_salary')) if row.get('min_salary') else None,
        max_salary  = _decimal(row.get('max_salary')) if row.get('max_salary') else None,
        location    = loc,
        views_count = _int(row.get('views_count')),
        is_active   = True,
    )
    listing.save()
    listing.set_current_language('ar')
    listing.title       = _str(row.get('title')) or f'وظيفة {post_id}'
    listing.description = _str(row.get('description'))
    listing.save()


def seed_jobs(rows: list[dict], user_map: dict, loc_cache: dict,
              dry_run: bool, stdout) -> int:
    if not user_map:
        stdout.write("  [WARN] user_map is empty — all job rows will be skipped. Seed users first.")
        return 0

    total = len(rows)
    stdout.write(f"  {total} job rows to process ...")

    if dry_run:
        would_create = sum(1 for r in rows if user_map.get(_str(r.get('phone'))))
        stdout.write(f"  JobListing done: {would_create} would create, {total - would_create} no-user")
        return would_create

    created = skipped = errors = 0
    for idx, row in enumerate(rows, 1):
        if not user_map.get(_str(row.get('phone'))):
            skipped += 1
            continue
        post_id = _str(row.get('source_post_id'))
        try:
            with transaction.atomic():
                _save_job_row(row, user_map, loc_cache)
            created += 1
        except Exception as exc:
            errors += 1
            stdout.write(f"  [job error] {post_id}: {exc}")
        _progress(stdout, "job", idx, total)

    stdout.write(f"  JobListing done: {created} created, {skipped} no-user, {errors} errors")
    return created


# ── Command ────────────────────────────────────────────────────────────────────

class Command(BaseCommand):
    help = 'Seed database from facebook/seed_data.xlsx (generated by extract_to_excel.py)'

    def add_arguments(self, parser):
        parser.add_argument(
            '--excel',
            default='../facebook/seed_data.xlsx',
            help='Path to the Excel file (default: ../facebook/seed_data.xlsx)',
        )
        parser.add_argument(
            '--sheet',
            default='all',
            choices=['all', 'users', 'PropertyListing', 'VehicleListing', 'ClassifiedListing', 'JobListing'],
            help='Which sheet to seed (default: all)',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Parse and validate without writing to the database',
        )
        parser.add_argument(
            '--skip-images',
            action='store_true',
            help='Skip image uploads (much faster; images can be added later)',
        )
        parser.add_argument(
            '--limit',
            type=int,
            default=None,
            help='Max rows to load per sheet, including users (e.g. --limit 50 for testing)',
        )

    def handle(self, *_args, **options):
        excel_path = Path(options['excel'])
        if not excel_path.exists():
            raise CommandError(
                f"Excel file not found: {excel_path}\n"
                "Generate it first:  python ../facebook/extract_to_excel.py"
            )

        dry_run     = options['dry_run']
        skip_images = options['skip_images']
        sheet       = options['sheet']
        limit       = options['limit']

        self.stdout.write(
            self.style.SUCCESS(
                f"Loading {excel_path}  (dry_run={dry_run}, skip_images={skip_images})"
            )
        )

        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

        def get_rows(name: str) -> list[dict]:
            if name not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f"  Sheet '{name}' not found, skipping."))
                return []
            rows = _read_sheet(wb[name])
            if limit is not None:
                rows = rows[:limit]
            self.stdout.write(f"  Sheet '{name}': {len(rows)} rows" + (f" (limited to {limit})" if limit else ""))
            return rows

        # Pre-load location cache (one query for all governorates)
        self.stdout.write("Pre-loading location cache ...")
        loc_cache = _build_location_cache()
        self.stdout.write(f"  {len(loc_cache)} governorates cached")

        # Users
        user_map: dict = {}
        if sheet in ('all', 'users'):
            self.stdout.write(self.style.HTTP_INFO("Seeding users ..."))
            user_map = seed_users(get_rows('users'), dry_run, self.stdout)
        else:
            from apps.accounts.models import User
            self.stdout.write("Building user map from existing DB records ...")
            for u in User.objects.filter(phone__gt='').only('id', 'phone'):
                user_map[u.phone] = u
            self.stdout.write(f"  {len(user_map)} users loaded from DB")

        # Listings
        if sheet in ('all', 'PropertyListing'):
            self.stdout.write(self.style.HTTP_INFO("Seeding PropertyListing ..."))
            seed_properties(get_rows('PropertyListing'), user_map, loc_cache, dry_run, skip_images, self.stdout)

        if sheet in ('all', 'VehicleListing'):
            self.stdout.write(self.style.HTTP_INFO("Seeding VehicleListing ..."))
            seed_vehicles(get_rows('VehicleListing'), user_map, loc_cache, dry_run, skip_images, self.stdout)

        if sheet in ('all', 'ClassifiedListing'):
            self.stdout.write(self.style.HTTP_INFO("Seeding ClassifiedListing ..."))
            seed_classifieds(get_rows('ClassifiedListing'), user_map, loc_cache, dry_run, skip_images, self.stdout)

        if sheet in ('all', 'JobListing'):
            self.stdout.write(self.style.HTTP_INFO("Seeding JobListing ..."))
            seed_jobs(get_rows('JobListing'), user_map, loc_cache, dry_run, self.stdout)

        if dry_run:
            self.stdout.write(self.style.WARNING("\nDRY RUN complete — no changes committed."))
        else:
            self.stdout.write(self.style.SUCCESS("\nDone! All records committed."))
